# ── Cargar variables de entorno desde .env (si existe) ────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv no instalado, se usarán las vars del sistema

from flask import Flask, request, jsonify, send_from_directory, session, render_template, redirect, url_for
from werkzeug.security import check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
import os
import uuid
import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
import database
import io
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta
from datetime import datetime, timedelta

import sys

# ── Módulo de licencias Supabase ───────────────────────────────────────────────
try:
    import supabase_license as _lic
    LICENSE_MODULE_OK = True
except ImportError:
    LICENSE_MODULE_OK = False

# ── Mercado Pago SDK ───────────────────────────────────────────────────────────
try:
    import mercadopago
    MP_AVAILABLE = True
except ImportError:
    MP_AVAILABLE = False

# ===========================
# CONTRASEÑA MAESTRA (Cargar desde .env por seguridad)
# ===========================
MASTER_PASSWORD = os.environ.get("MASTER_PASSWORD", "Minecra32")

# ── Configuración de Mercado Pago ─────────────────────────────────────────────
MP_ACCESS_TOKEN = os.environ.get("MERCADO_PAGO_ACCESS_TOKEN", "")
APP_BASE_URL    = os.environ.get("APP_BASE_URL", "http://localhost:5000")
PRECIO_MENSUAL  = 60000  # COP — Mensualidad Mordev POS

def get_base_path():
    """Obtiene la ruta base del proyecto, compatible con ejecutables de PyInstaller."""
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def get_external_data_path():
    """Obtiene la ruta para datos externos persistentes (BD, fotos), fuera del bundle temporal."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__, 
            static_folder=os.path.join(get_base_path(), 'frontend'), 
            static_url_path='')
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'clave-secreta-ventas-app-2026')

# Carpeta de templates para las pantallas de licencia
app.template_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates')

# Carpeta para subir archivos (en la misma carpeta que el ejecutable si está empaquetado)
UPLOAD_FOLDER = os.path.join(get_external_data_path(), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

def allowed_image(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS


# --- Decoradores de autenticación ---

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session or 'business_id' not in session:
            session.clear()
            return jsonify({"error": "No autorizado"}), 401
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({"error": "No autenticado"}), 401
        if session.get('role') != 'admin':
            return jsonify({"error": "Acceso denegado. Se requiere rol de administrador"}), 403
        return f(*args, **kwargs)
    return decorated


# ─────────────────────────────────────────────────────────────────────────────
# DECORADOR DE LICENCIA — verifica Supabase en cada carga de página HTML
# Solo se aplica en rutas que sirven páginas (no en endpoints de API JSON)
# ─────────────────────────────────────────────────────────────────────────────
def requiere_licencia(f):
    """Middleware: bloquea el acceso si la licencia del negocio está vencida."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not LICENSE_MODULE_OK:
            return f(*args, **kwargs)
            
        business_id = session.get('business_id')
        if not business_id:
            return f(*args, **kwargs) # Si no hay sesión, dejamos que @login_required lo maneje
            
        info = _lic.get_license_info(business_id)
        if not info.get('activa', True):
            return jsonify({"error": "Licencia vencida", "vencimiento": info.get('fecha_vencimiento')}), 403
        return f(*args, **kwargs)
    return decorated


# --- Rutas del Frontend ---

@app.before_request
def log_request_info():
    print(f"[REQUEST] {request.method} {request.path}")

@app.route('/')
def serve_landing():
    # Si ya tiene sesión, lo mandamos al POS directamente
    if 'business_id' in session:
        return redirect('/pos')
    
    print(f"[DEBUG] Static folder: {app.static_folder}")
    full_path = os.path.join(app.static_folder, 'landing.html')
    print(f"[DEBUG] Full path to landing: {full_path}")
    print(f"[DEBUG] File exists: {os.path.exists(full_path)}")
    
    return send_from_directory(app.static_folder, 'landing.html')

@app.route('/pos')
def serve_index():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/master')
def serve_master():
    return send_from_directory('../frontend', 'master.html')

@app.route('/socio')
def serve_socio():
    return send_from_directory('../frontend', 'socio.html')

@app.route('/register')
def serve_register():
    return send_from_directory(app.static_folder, 'register.html')

@app.route('/restablecer')
def serve_restablecer():
    """Página de restablecimiento de contraseña (recibe ?token=...)"""
    return send_from_directory(app.static_folder, 'restablecer.html')


@app.route('/uploads/<filename>')
def serve_upload(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


# --- API de Autenticación y Registro ---

@app.route('/api/register', methods=['POST'])
def register_business():
    data = request.get_json()
    if not data or 'nombre_negocio' not in data or 'email' not in data or 'password' not in data:
        return jsonify({"error": "Faltan campos obligatorios"}), 400

    # Manejo seguro de campos opcionales
    codigo_ref = data.get('codigo_referido')
    if codigo_ref:
        codigo_ref = str(codigo_ref).strip() or None

    result = database.registrar_nuevo_negocio(
        data['nombre_negocio'],
        data['email'],
        data['password'],
        codigo_referido=codigo_ref,
        categoria=data.get('categoria', 'general'),
        color_hex=data.get('color_hex'),
        icono_slug=data.get('icono_slug'),
    )
    if 'error' in result:
        return jsonify(result), 400

    # Enviar correo de bienvenida (no bloquea si falla)
    try:
        from email_service import send_welcome_email
        send_welcome_email(data['email'], data['nombre_negocio'], data.get('categoria', 'general'))
    except Exception as e:
        print(f"[REGISTER] Email bienvenida falló: {e}")

    return jsonify(result), 201


# ─────────────────────────────────────────────────────────────────────────────
# DESIGN TOKENS — Tema visual del negocio
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/business/theme', methods=['GET'])
def get_theme():
    """Devuelve el tema visual del negocio autenticado (público para splash screen)."""
    business_id = session.get('business_id')
    if not business_id:
        return jsonify(database.THEME_PRESETS['general'])
    theme = database.get_business_theme(business_id)
    return jsonify(theme)


@app.route('/api/business/theme', methods=['PUT'])
@admin_required
def update_theme():
    """Actualiza el tema visual del negocio. Solo admin."""
    data = request.get_json()
    if not data or 'categoria' not in data:
        return jsonify({"error": "Falta el campo 'categoria'"}), 400

    categorias_validas = ['mascotas', 'carros', 'comida', 'tecnologia', 'general']
    if data['categoria'] not in categorias_validas:
        return jsonify({"error": f"Categoría inválida. Opciones: {categorias_validas}"}), 400

    result = database.update_business_theme(
        session['business_id'],
        data['categoria'],
        data.get('color_hex'),
        data.get('icono_slug'),
    )
    if 'error' in result:
        return jsonify(result), 400
    return jsonify(result)


# ─────────────────────────────────────────────────────────────────────────────
# RESET PASSWORD — Flujo seguro con Resend + token de 3 minutos
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/password-reset/request', methods=['POST'])
def request_password_reset():
    """Solicita reset de contraseña. Genera token y envía email via Resend."""
    data = request.get_json()
    if not data or 'email' not in data:
        return jsonify({"error": "Falta el campo email"}), 400

    email = data['email'].strip().lower()
    result = database.create_password_reset_token(email)

    if 'error' in result:
        return jsonify(result), 500

    # Solo enviar email si se generó token real (el email existe)
    if result.get('raw_token'):
        try:
            from email_service import send_password_reset_email
            reset_url = f"{APP_BASE_URL}/restablecer?token={result['raw_token']}"
            email_res = send_password_reset_email(email, result.get('nombre_negocio', ''), reset_url)
            
            if 'error' in email_res:
                print(f"[RESET] Error de Resend: {email_res['error']}")
            else:
                print(f"[RESET] Correo enviado exitosamente a {email}")
                
        except Exception as e:
            print(f"[RESET] Excepción fatal enviando email: {str(e)}")

    # Siempre retornar el mismo mensaje (no revelar si el email existe)
    return jsonify({"success": True, "message": "Si el correo existe, recibirás las instrucciones en unos momentos."})


@app.route('/api/password-reset/verify', methods=['POST'])
def verify_password_reset():
    """Valida el token y actualiza la contraseña."""
    data = request.get_json()
    if not data or 'token' not in data or 'new_password' not in data:
        return jsonify({"error": "Faltan campos: token, new_password"}), 400

    if len(data['new_password']) < 6:
        return jsonify({"error": "La contraseña debe tener al menos 6 caracteres"}), 400

    # Validar token
    validation = database.validate_reset_token(data['token'])
    if not validation.get('valid'):
        return jsonify({"error": validation.get('reason', 'Token inválido')}), 400

    # Actualizar contraseña
    result = database.update_password_by_email(
        validation['email'],
        data['new_password'],
        token_id=validation.get('token_id')
    )
    if 'error' in result:
        return jsonify(result), 500

    return jsonify({"success": True, "message": "Contraseña actualizada correctamente. Ya puedes iniciar sesión."})


# ─────────────────────────────────────────────────────────────────────────────
# VENDEDORES (socios comerciales)
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/vendedores/verificar/<codigo>', methods=['GET'])
def verificar_codigo_referido(codigo):
    """Verifica si un código de referido es válido (usado en el registro)."""
    vend = database.get_vendedor_by_codigo(codigo.upper().strip())
    if vend:
        return jsonify({"valido": True, "nombre_vendedor": vend['nombre']})
    return jsonify({"valido": False})


# ─────────────────────────────────────────────────────────────────────────────
# LIQUIDACIONES — Panel admin (requiere admin)
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/admin/liquidaciones', methods=['GET'])
@admin_required
def get_liquidaciones():
    """Devuelve la liquidación mensual de vendedores. Solo admin."""
    data = database.get_liquidacion_vendedores()
    return jsonify(data)


@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    if not data or 'username' not in data or 'password' not in data:
        return jsonify({"error": "Faltan usuario o contraseña"}), 400

    user = database.get_user_by_username_and_business(data['username'])
    if not user or not check_password_hash(user['password_hash'], data['password']):
        return jsonify({"error": "Usuario o contraseña incorrectos"}), 401

    session['user_id'] = user['id']
    session['username'] = user['username']
    session['role'] = user['role']
    session['business_id'] = user['id_negocio']

    return jsonify({
        "success": True,
        "user": {
            "id": user['id'],
            "username": user['username'],
            "role": user['role'],
            "avatar_path": user.get('avatar_path', '')
        }
    })

# ── MASTER ADMIN (Solo Desarrollador) ────────────────────────────────────────
def check_master_auth():
    return request.headers.get('x-master-key') == 'minecra32'

@app.route('/api/master/vendedores', methods=['GET', 'POST'])
def master_vendedores():
    if not check_master_auth():
        return jsonify({"error": "No autorizado"}), 401
    
    client = database.get_client()
    if request.method == 'GET':
        res = client.table("vendedores").select("*").order("created_at", desc=True).execute()
        return jsonify(res.data)
    
    if request.method == 'POST':
        data = request.get_json()
        import random, string
        codigo = data.get('codigo_referido') or ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
        
        nuevo_vendedor = {
            "nombre": data['nombre'],
            "email": data.get('email'),
            "codigo_referido": codigo.upper(),
            "password": data.get('password'), # Añadido campo password
            "comision_porcentaje": 0.20,
            "datos_pago": data.get('datos_pago', {}),
            "activo": True
        }
        res = client.table("vendedores").insert(nuevo_vendedor).execute()
        return jsonify(res.data[0])

@app.route('/api/master/vendedores/<id>', methods=['PUT'])
def master_update_vendedor(id):
    if not check_master_auth():
        return jsonify({"error": "No autorizado"}), 401
        
    client = database.get_client()
    data = request.get_json()
    res = client.table("vendedores").update(data).eq("id", id).execute()
    return jsonify(res.data[0])

@app.route('/api/master/liquidaciones', methods=['GET'])
def master_liquidaciones():
    if not check_master_auth():
        return jsonify({"error": "No autorizado"}), 401
        
    client = database.get_client()
    res = client.table("liquidacion_vendedores").select("*").execute()
    return jsonify(res.data)

# ── PORTAL DE SOCIOS (Vendedores) ─────────────────────────────────────────────
@app.route('/api/vendedor/stats', methods=['POST'])
def vendedor_stats():
    data = request.get_json()
    codigo = data.get('codigo')
    password = data.get('password')
    
    client = database.get_client()
    # 1. Buscar vendedor con código y contraseña
    vend = client.table("vendedores").select("*").eq("codigo_referido", codigo).eq("password", password).eq("activo", True).maybe_single().execute()
    
    if not vend.data:
        return jsonify({"error": "Código o contraseña incorrectos"}), 404
    
    # 2. Obtener sus estadísticas de la vista
    stats = client.table("liquidacion_vendedores").select("*").eq("codigo_referido", codigo).maybe_single().execute()
    
    # 3. Obtener lista de negocios traídos
    negocios = client.table("negocios").select("nombre_negocio, created_at, licencia_activa").eq("vendedor_id", vend.data['id']).execute()
    
    return jsonify({
        "vendedor": vend.data,
        "stats": stats.data or {"negocios_activos_mes": 0, "comision_a_pagar": 0},
        "negocios": negocios.data
    })


@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({"success": True})


@app.route('/api/me', methods=['GET'])
@login_required
def get_current_user():
    user = database.get_user_by_username_and_business(session['username'], session['business_id'])
    
    # Obtener información de licencia para mostrar alertas (ej. 5 días)
    lic_info = {}
    if LICENSE_MODULE_OK:
        lic_info = _lic.get_license_info(session['business_id'])
        
    return jsonify({
        "id": session['user_id'],
        "username": session['username'],
        "role": session['role'],
        "avatar_path": user.get('avatar_path', '') if user else '',
        "license": lic_info
    })


# --- API de Usuarios ---

@app.route('/api/users', methods=['GET'])
@admin_required
def list_users():
    users = database.get_all_users(session.get('business_id'))
    return jsonify(users)


@app.route('/api/users', methods=['POST'])
@admin_required
def create_user():
    data = request.get_json()
    if not data or 'username' not in data or 'password' not in data or 'role' not in data:
        return jsonify({"error": "Faltan campos: username, password, role"}), 400

    result = database.create_user(session.get('business_id'), data['username'], data['password'], data['role'])
    if 'error' in result:
        return jsonify(result), 400
    return jsonify(result), 201


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def remove_user(user_id):
    result = database.delete_user(session.get('business_id'), user_id)
    if 'error' in result:
        return jsonify(result), 400
    return jsonify(result)


@app.route('/api/users/<int:user_id>/avatar', methods=['POST'])
@login_required
def upload_avatar(user_id):
    # Only admin or the user themselves can change avatar
    if session['user_id'] != user_id and session.get('role') != 'admin':
        return jsonify({"error": "Acceso denegado"}), 403

    if 'file' not in request.files:
        return jsonify({"error": "No se envió ningún archivo"}), 400

    file = request.files['file']
    if file.filename == '' or not allowed_image(file.filename):
        return jsonify({"error": "Formato de imagen no válido. Use: png, jpg, jpeg, gif, webp"}), 400

    ext = file.filename.rsplit('.', 1)[1].lower()
    filename = f"avatar_{user_id}_{uuid.uuid4().hex[:8]}.{ext}"
    
    # Subir a Supabase en lugar de local
    avatar_url = database.upload_file_to_supabase(file, filename)
    
    if not avatar_url:
        return jsonify({"error": "Error al subir la imagen a la nube"}), 500

    database.update_user_avatar(session.get('business_id'), user_id, avatar_url)

    return jsonify({"success": True, "avatar_path": avatar_url})


# --- API de Productos ---

@app.route('/api/products', methods=['GET'])
@login_required
def list_products():
    products = database.get_all_products(session.get('business_id'))
    return jsonify(products)


@app.route('/api/products/<int:product_id>', methods=['GET'])
@login_required
def get_product(product_id):
    product = database.get_product_by_id(session.get('business_id'), product_id)
    if not product:
        return jsonify({"error": "Producto no encontrado"}), 404
    return jsonify(product)


@app.route('/api/products', methods=['POST'])
@admin_required
def add_product():
    # Helper para evitar ValueErrors si viene vacío
    def safe_float(val):
        if val in (None, '', 'NaN', 'nan'): return 0.0
        try: return float(val)
        except ValueError: return 0.0
    def safe_int(val):
        if val in (None, '', 'NaN', 'nan'): return 0
        try: return int(float(val))
        except ValueError: return 0

    # Support both JSON and multipart form
    if request.content_type and 'multipart/form-data' in request.content_type:
        name = request.form.get('name')
        price = request.form.get('price', 0)
        if not name:
            return jsonify({"error": "Faltan campos: name"}), 400

        image_path = ''
        if 'image' in request.files:
            file = request.files['image']
            if file.filename and allowed_image(file.filename):
                ext = file.filename.rsplit('.', 1)[1].lower()
                filename = f"product_{uuid.uuid4().hex[:8]}.{ext}"
                image_path = database.upload_file_to_supabase(file, filename)

        try:
            product_id = database.add_product(session.get('business_id'), 
                name=name,
                category=request.form.get('category', ''),
                price=safe_float(price),
                stock=safe_int(request.form.get('stock')),
                image_path=image_path,
                reference=request.form.get('reference', ''),
                unit=request.form.get('unit', ''),
                purchase_price=safe_float(request.form.get('purchase_price')),
                sale_price=safe_float(request.form.get('sale_price', price)),
                is_bulk=safe_int(request.form.get('is_bulk')),
                barcode=request.form.get('barcode') or None
            )
        except Exception as e:
            return jsonify({"error": str(e)}), 400
    else:
        data = request.get_json()
        if not data or 'name' not in data:
            return jsonify({"error": "Faltan campos: name"}), 400
        
        try:
            price = data.get('price', 0)
            product_id = database.add_product(session.get('business_id'), 
                name=data['name'],
                category=data.get('category', ''),
                price=safe_float(price),
                stock=safe_int(data.get('stock')),
                image_path=data.get('image_path', ''),
                reference=data.get('reference', ''),
                unit=data.get('unit', ''),
                purchase_price=safe_float(data.get('purchase_price')),
                sale_price=safe_float(data.get('sale_price', price)),
                is_bulk=safe_int(data.get('is_bulk')),
                barcode=data.get('barcode') or None
            )
        except Exception as e:
            return jsonify({"error": str(e)}), 400
    return jsonify({"success": True, "id": product_id}), 201


@app.route('/api/products/<int:product_id>', methods=['PUT'])
@admin_required
def update_product(product_id):
    existing = database.get_product_by_id(session.get('business_id'), product_id)
    if not existing:
        return jsonify({"error": "Producto no encontrado"}), 404

    if request.content_type and 'multipart/form-data' in request.content_type:
        image_path = existing.get('image_path')
        if 'image' in request.files:
            file = request.files['image']
            if file.filename and allowed_image(file.filename):
                ext = file.filename.rsplit('.', 1)[1].lower()
                filename = f"product_{product_id}_{uuid.uuid4().hex[:8]}.{ext}"
                image_path = database.upload_file_to_supabase(file, filename)

        database.update_product(session.get('business_id'), 
            product_id=product_id,
            name=request.form.get('name'),
            category=request.form.get('category'),
            price=float(request.form.get('price')) if request.form.get('price') is not None else None,
            stock=int(request.form.get('stock')) if request.form.get('stock') is not None else None,
            image_path=image_path,
            reference=request.form.get('reference'),
            unit=request.form.get('unit'),
            purchase_price=float(request.form.get('purchase_price')) if request.form.get('purchase_price') is not None else None,
            sale_price=float(request.form.get('sale_price')) if request.form.get('sale_price') is not None else None,
            is_bulk=int(request.form.get('is_bulk', 0)) if request.form.get('is_bulk') is not None else None,
            barcode=request.form.get('barcode') or None
        )
    else:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No se enviaron datos"}), 400
        database.update_product(session.get('business_id'), 
            product_id=product_id,
            name=data.get('name'),
            category=data.get('category'),
            price=float(data['price']) if 'price' in data else None,
            stock=int(data['stock']) if 'stock' in data else None,
            image_path=data.get('image_path'),
            reference=data.get('reference'),
            unit=data.get('unit'),
            purchase_price=float(data['purchase_price']) if 'purchase_price' in data else None,
            sale_price=float(data['sale_price']) if 'sale_price' in data else None,
            is_bulk=int(data['is_bulk']) if 'is_bulk' in data else None,
            barcode=data.get('barcode') or None
        )
    return jsonify({"success": True})


@app.route('/api/products/<int:product_id>', methods=['DELETE'])
@admin_required
def remove_product(product_id):
    existing = database.get_product_by_id(session.get('business_id'), product_id)
    if not existing:
        return jsonify({"error": "Producto no encontrado"}), 404
    database.delete_product(session.get('business_id'), product_id)
    return jsonify({"success": True})


# --- API Búsqueda por Código de Barras ---

@app.route('/api/products/barcode/<barcode>', methods=['GET'])
@login_required
def get_product_by_barcode(barcode):
    """Busca un producto por su código de barras (campo barcode) o referencia."""
    product = database.get_product_by_barcode(session.get('business_id'), barcode)
    if not product:
        return jsonify({"error": "Producto no encontrado"}), 404
    return jsonify(product)


# --- API de Exportar Inventario Excel ---

@app.route('/api/products/export-excel', methods=['GET'])
@admin_required
def export_inventory_excel():
    products = database.get_all_products(session.get('business_id'))

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = ["ID", "Nombre", "Referencia", "Unidad", "Categoría", "P. Compra (COP)", "P. Venta (COP)", "Stock"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for p in products:
        ws.append([
            p['id'],
            p['name'],
            p.get('reference', ''),
            p.get('unit', ''),
            p.get('category', ''),
            p.get('purchase_price', 0),
            p.get('sale_price') or p.get('price', 0),
            p['stock']
        ])

    wb.save(output)
    output.seek(0)
    from datetime import datetime, timedelta
    filename = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# --- API de Importar Excel ---

@app.route('/api/products/import-excel', methods=['POST'])
@admin_required
def import_excel():
    if 'file' not in request.files:
        return jsonify({"error": "No se envió ningún archivo"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No se seleccionó ningún archivo"}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "El archivo debe ser .xlsx o .xls"}), 400

    filepath = os.path.join(UPLOAD_FOLDER, 'import_temp.xlsx')
    file.save(filepath)

    result = database.import_from_excel(session.get('business_id'), filepath)

    # Limpiar archivo temporal
    try:
        os.remove(filepath)
    except:
        pass

    if 'error' in result:
        return jsonify(result), 400
    return jsonify(result)


# --- API de Ventas ---

@app.route('/api/sales', methods=['POST'])
@login_required
def create_sale():
    data = request.get_json()
    if not data or 'product_id' not in data or 'quantity' not in data:
        return jsonify({"error": "Faltan campos: product_id, quantity"}), 400

    result = database.register_sale(
        session.get('business_id'),
        product_id=int(data['product_id']),
        quantity=int(data['quantity']),
        seller_id=session['user_id']
    )

    if 'error' in result:
        return jsonify(result), 400
    return jsonify(result), 201


@app.route('/api/sales', methods=['GET'])
@login_required
def list_sales():
    if session.get('role') == 'admin':
        sales = database.get_sales(session.get('business_id'))
    else:
        sales = database.get_sales(session.get('business_id'), seller_id=session['user_id'])
    return jsonify(sales)


@app.route('/api/sales/<int:sale_id>', methods=['DELETE'])
@admin_required
def delete_sale(sale_id):
    result = database.delete_sale(session.get('business_id'), sale_id)
    if 'error' in result:
        return jsonify(result), 404
    return jsonify(result)


# --- API de Clientes ---

@app.route('/api/customers', methods=['GET'])
@login_required
def list_customers():
    customers = database.get_all_customers(session.get('business_id'))
    return jsonify(customers)


@app.route('/api/customers/by-nid/<nid>', methods=['GET'])
@login_required
def get_customer_by_nid(nid):
    customer = database.get_customer_by_nid(session.get('business_id'), nid)
    if not customer:
        return jsonify({"error": "Cliente no encontrado"}), 404
    return jsonify(customer)


@app.route('/api/customers', methods=['POST'])
@login_required
def create_customer():
    data = request.get_json()
    if not data or 'name' not in data:
        return jsonify({"error": "Falta el nombre del cliente"}), 400
    
    result = database.create_customer(session.get('business_id'), 
        name=data['name'],
        address=data.get('address', ''),
        phone=data.get('phone', ''),
        nid=data.get('nid', ''),
        placa=data.get('placa', ''),
        vehiculo=data.get('vehiculo', '')
    )
    if 'error' in result:
        return jsonify(result), 400
    return jsonify(result), 201


# --- API de Facturación ---

@app.route('/api/invoices', methods=['POST'])
@login_required
def create_invoice():
    data = request.get_json()
    if not data or 'items' not in data or 'customer' not in data or 'payment' not in data:
        return jsonify({"error": "Faltan datos de la factura (items, cliente o pago)"}), 400

    result = database.create_invoice(session.get('business_id'), 
        items=data['items'],
        customer_data=data['customer'],
        payment_info=data['payment'],
        seller_id=session['user_id'],
        draft_id=data.get('draft_id')
    )

    if 'error' in result:
        return jsonify(result), 400
    return jsonify(result), 201


@app.route('/api/invoices/<int:invoice_id>', methods=['GET'])
@login_required
def get_invoice(invoice_id):
    invoice = database.get_invoice_details(session.get('business_id'), invoice_id)
    if not invoice:
        return jsonify({"error": "Factura no encontrada"}), 404
    return jsonify(invoice)


@app.route('/api/invoices/<int:invoice_id>', methods=['DELETE'])
@admin_required
def delete_invoice_api(invoice_id):
    result = database.delete_invoice(session.get('business_id'), invoice_id)
    if 'error' in result:
        return jsonify(result), 404
    return jsonify(result)


# --- API de Borradores (Facturas Abiertas) ---

@app.route('/api/drafts', methods=['POST'])
@login_required
def save_draft():
    data = request.get_json()
    if not data or 'items' not in data or 'customer' not in data or 'payment' not in data:
        return jsonify({"error": "Faltan datos del borrador"}), 400

    result = database.save_draft(session.get('business_id'), 
        items=data['items'],
        customer_data=data['customer'],
        payment_info=data['payment'],
        seller_id=session['user_id'],
        draft_id=data.get('draft_id')
    )

    if 'error' in result:
        return jsonify(result), 400
    return jsonify(result), 201


@app.route('/api/drafts', methods=['GET'])
@login_required
def list_drafts():
    if session.get('role') == 'admin':
        drafts = database.get_all_drafts(session.get('business_id'))
    else:
        drafts = database.get_all_drafts(session.get('business_id'), seller_id=session['user_id'])
    return jsonify(drafts)


@app.route('/api/drafts/<int:draft_id>', methods=['GET'])
@login_required
def get_draft(draft_id):
    draft = database.get_draft_details(session.get('business_id'), draft_id)
    if not draft:
        return jsonify({"error": "Borrador no encontrado"}), 404
    return jsonify(draft)


@app.route('/api/drafts/<int:draft_id>', methods=['DELETE'])
@login_required
def remove_draft(draft_id):
    result = database.delete_draft(session.get('business_id'), draft_id)
    return jsonify(result)



@app.route('/api/reports/sales', methods=['GET'])
@admin_required
def export_sales_report():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    report_format = request.args.get('format', 'excel').lower()

    if not start_date or not end_date:
        return jsonify({"error": "Faltan fechas de inicio y fin"}), 400

    sales = database.get_sales_by_date_range(session.get('business_id'), start_date, end_date)
    
    if report_format == 'excel':
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas"

        # Encabezados
        headers = ["ID", "Producto", "Cantidad", "P. Unit (COP)", "P. Coste Unit (COP)", "Total Venta (COP)", "Total Coste (COP)", "Ganancia (COP)", "Vendedor", "Fecha"]
        ws.append(headers)
        
        # Estilos encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        total_venta = 0
        total_coste = 0
        for s in sales:
            coste_total = (s.get('purchase_unit_price') or 0) * s['quantity']
            ganancia = s['total'] - coste_total
            ws.append([
                s['id'], 
                s['product_name'], 
                s['quantity'], 
                s['unit_price'], 
                s.get('purchase_unit_price') or 0,
                s['total'],
                coste_total,
                ganancia,
                s['seller_name'], 
                s['date']
            ])
            total_venta += s['total']
            total_coste += coste_total

        ws.append([])
        ws.append(["", "", "", "", "", "TOTAL VENTAS", total_venta])
        ws.append(["", "", "", "", "", "TOTAL COSTE", total_coste])
        ws.append(["", "", "", "", "", "GANANCIA TOTAL", total_venta - total_coste])
        
        # Estilos finales
        ws.cell(row=ws.max_row-2, column=7).font = Font(bold=True)
        ws.cell(row=ws.max_row-1, column=7).font = Font(bold=True)
        last_total_cell = ws.cell(row=ws.max_row, column=7)
        last_total_cell.font = Font(bold=True, color="008000") # Verde para ganancia

        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'reporte_ventas_{start_date}_a_{end_date}.xlsx'
        )

    return jsonify({"error": "Formato no soportado"}), 400





# --- API de Báscula ---

@app.route('/api/scale/ports', methods=['GET'])
@login_required
def list_scale_ports():
    """Lista los puertos COM disponibles."""
    try:
        import serial.tools.list_ports
        ports = [{'port': p.device, 'description': p.description} 
                 for p in serial.tools.list_ports.comports()]
        return jsonify({"ports": ports})
    except ImportError:
        return jsonify({"error": "pyserial no está instalado", "ports": []})


@app.route('/api/scale/config', methods=['GET'])
@login_required
def get_scale_config():
    """Retorna la configuración actual de la báscula."""
    return jsonify({
        "port": database.get_setting(session.get('business_id'), 'scale_port', ''),
        "baudrate": database.get_setting(session.get('business_id'), 'scale_baudrate', '9600'),
        "protocol": database.get_setting(session.get('business_id'), 'scale_protocol', 'generic')
    })


@app.route('/api/scale/config', methods=['POST'])
@admin_required
def save_scale_config():
    """Guarda la configuración de la báscula."""
    data = request.get_json()
    if data.get('port'): database.set_setting(session.get('business_id'), 'scale_port', data['port'])
    if data.get('baudrate'): database.set_setting(session.get('business_id'), 'scale_baudrate', data['baudrate'])
    if data.get('protocol'): database.set_setting(session.get('business_id'), 'scale_protocol', data['protocol'])
    return jsonify({"success": True})


@app.route('/api/scale/read', methods=['GET'])
@login_required
def read_scale_weight():
    """Lee el peso de la báscula via puerto serial."""
    port = database.get_setting(session.get('business_id'), 'scale_port', '')
    baudrate = int(database.get_setting(session.get('business_id'), 'scale_baudrate', '9600'))
    
    if not port:
        return jsonify({"error": "Báscula no configurada"}), 400
    
    try:
        import serial
        with serial.Serial(port, baudrate, timeout=2) as ser:
            # Intentar leer peso - protocolos genéricos
            # Vacíamos buffer primero
            ser.reset_input_buffer()
            # Enviar comando de lectura (común en muchas básculas genéricas)
            ser.write(b'\r\n')
            raw = ser.readline().decode('ascii', errors='ignore').strip()
            
            if not raw:
                # Intentar leer sin comando
                raw = ser.read(32).decode('ascii', errors='ignore').strip()
            
            # Extraer el valor numérico del string
            import re
            # Buscar patron numérico con decimales (ej: 1.234, +001.23, 1234g, etc.)
            match = re.search(r'[+-]?\s*([0-9]+[.,][0-9]+|[0-9]+)', raw)
            if match:
                weight_str = match.group(1).replace(',', '.')
                weight = float(weight_str)
                return jsonify({"weight": weight, "raw": raw, "unit": "kg"})
            else:
                return jsonify({"error": f"Respuesta no reconocida: {raw}", "raw": raw}), 422
    except ImportError:
        return jsonify({"error": "pyserial no instalado. Ejecuta: pip install pyserial"}), 500
    except Exception as e:
        return jsonify({"error": f"Error al leer báscula: {str(e)}"}), 500


# ─────────────────────────────────────────────────────────────────────────────
# RUTAS DE LICENCIA Y PAGOS — Mercado Pago + Supabase
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/licencia/estado', methods=['GET'])
def licencia_estado():
    """Retorna el estado actual de la licencia (para mostrar en el POS)."""
    if not LICENSE_MODULE_OK:
        return jsonify({"activa": False, "dias_restantes": 0, "error": "Módulo no instalado"})
    try:
        info = _lic.get_license_info(session.get('business_id'))
        fv = info.get('fecha_vencimiento')
        return jsonify({
            "activa": info.get('activa', False),
            "dias_restantes": info.get('dias_restantes', 0),
            "fecha_vencimiento": fv.strftime('%d/%m/%Y') if fv and hasattr(fv, 'strftime') else None,
            "error": info.get('error')
        })
    except Exception as e:
        return jsonify({"activa": False, "dias_restantes": 0, "error": str(e)}), 500


@app.route('/api/create-preference', methods=['POST'])
def create_preference():
    """
    Crea una preferencia de pago en Mercado Pago para la mensualidad.
    Retorna el init_point (URL de pago) al frontend.
    """
    if not MP_AVAILABLE:
        return jsonify({"error": "SDK de Mercado Pago no instalado. Ejecuta: pip install mercadopago"}), 500

    if not MP_ACCESS_TOKEN:
        return jsonify({"error": "MERCADO_PAGO_ACCESS_TOKEN no configurado"}), 500

    business_id = session.get('business_id')
    if not business_id:
        return jsonify({"error": "No hay sesión de negocio activa"}), 401

    try:
        sdk = mercadopago.SDK(MP_ACCESS_TOKEN)

        preference_data = {
            "items": [
                {
                    "id": "mordev-pos-mensual",
                    "title": "Mensualidad Mordev POS",
                    "description": "Suscripción mensual 30 días",
                    "category_id": "software",
                    "quantity": 1,
                    "currency_id": "COP",
                    "unit_price": float(PRECIO_MENSUAL),
                }
            ],
            "back_urls": {
                "success": f"{APP_BASE_URL}/",
                "failure": f"{APP_BASE_URL}/",
                "pending": f"{APP_BASE_URL}/",
            },
            "external_reference": business_id,
        }

        # Solo activar auto_return si tenemos HTTPS (Mercado Pago lo requiere)
        if "https" in APP_BASE_URL:
            preference_data["auto_return"] = "approved"

        # Solo enviar notification_url si no es localhost
        if "localhost" not in APP_BASE_URL:
            preference_data["notification_url"] = f"{APP_BASE_URL}/webhook-pagos"
        
        result = sdk.preference().create(preference_data)
        status = result.get("status")
        response_data = result.get("response", {})
        
        if status not in (200, 201):
            return jsonify({
                "error": f"Error de Mercado Pago (Status {status})", 
                "details": response_data,
                "raw": result # Para ver todo
            }), 500

        print(f"[OK MP] Preferencia creada exitosamente: {response_data.get('id')}")
        return jsonify({
            "init_point": response_data.get("init_point"),
            "preference_id": response_data.get("id"),
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/webhook-pagos', methods=['POST'])
def webhook_pagos():
    try:
        data = request.get_json(silent=True) or {}
        topic = data.get('type') or request.args.get('topic', '')
        payment_id = None

        if topic == 'payment':
            payment_id = str(data.get('data', {}).get('id', '') or request.args.get('id', ''))
        elif topic == 'merchant_order':
            return jsonify({"status": "ok"}), 200

        # --- MODO PRUEBA ---
        if request.args.get('test') == 'true':
            business_id = request.args.get('external_reference')
            if business_id:
                from supabase_license import acumular_dias
                acumular_dias(business_id, 30)
                print(f"[TEST] Simulación de pago exitosa para negocio {business_id}")
                return jsonify({"status": "ok", "msg": "Test mode processed"}), 200
        # -------------------

        if not payment_id or not MP_ACCESS_TOKEN:
            return jsonify({"status": "ok"}), 200

        sdk = mercadopago.SDK(MP_ACCESS_TOKEN)
        payment = sdk.payment().get(payment_id).get("response", {})
        
        if payment.get("status") == "approved":
            business_id = payment.get("external_reference")
            if business_id:
                # Usar el módulo de licencia para acumular días
                from supabase_license import acumular_dias
                result = acumular_dias(business_id, 30)
                print(f"[OK] Pago {payment_id} aprobado para negocio {business_id}")

        return jsonify({"status": "ok"}), 200

    except Exception as e:
        print(f"[ERROR] Error en webhook: {e}")
        return jsonify({"status": "ok"}), 200
        return jsonify({"status": "error", "msg": str(e)}), 200


# Rutas de retorno del pago (landing pages)
@app.route('/pago/exitoso')
def pago_exitoso():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/pago/fallido')
def pago_fallido():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/pago/pendiente')
def pago_pendiente():
    return send_from_directory(app.static_folder, 'index.html')


# --- Iniciar servidor ---

if __name__ == '__main__':
    import os
    port = int(os.environ.get("PORT", 5000))
    print(f"\nOK Servidor iniciado en el puerto {port}")
    print("   Modo Multi-cliente (SaaS) Activo")
    print("   Licencias:     Supabase", '[OK]' if LICENSE_MODULE_OK else '⚠️ módulo no disponible')
    print("   Mercado Pago:", '[OK] configurado' if MP_ACCESS_TOKEN else '⚠️ token no configurado', "\n")
    app.run(debug=True, host='0.0.0.0', port=port)
